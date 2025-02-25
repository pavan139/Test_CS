import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Load the Excel file
file_path = "/mnt/data/Incentive_Pay_Contribution_Audit1.xlsx"  # Update this with the actual file path
wb = load_workbook(file_path)
ws = wb["Summary"]  # Ensure this matches the sheet name in your file

# Given values for specific cells
data_values = {
    "B2": 100, "C2": 200,
    "B3": 150, "C3": 250,
    "B4": 180, "C4": 220,
    "B5": 130, "C5": 170
}

# Insert values into respective cells
for cell, value in data_values.items():
    ws[cell] = value

# Calculate and insert horizontal totals (D column)
for row in range(2, 6):
    ws[f"D{row}"] = ws[f"B{row}"].value + ws[f"C{row}"].value

# Calculate and insert vertical totals (6th row)
ws["B6"] = sum(ws[f"B{row}"].value for row in range(2, 6))
ws["C6"] = sum(ws[f"C{row}"].value for row in range(2, 6))
ws["D6"] = sum(ws[f"D{row}"].value for row in range(2, 6))

# Given DataFrame with Status Code, Status Description, and Values
df = pd.DataFrame({
    "Status Code": ["SC1", "SC2", "SC3"],
    "Status Description": ["Desc1", "Desc2", "Desc3"],
    "Value": [300, 400, 500]
})

# Border style definition
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Insert DataFrame values dynamically after row 12
start_row = 13
for index, row in df.iterrows():
    ws[f"A{start_row}"] = row["Status Code"]
    ws[f"B{start_row}"] = row["Status Description"]
    ws[f"C{start_row}"] = row["Value"]

    # Apply border to the inserted row
    for col in ["A", "B", "C"]:
        ws[f"{col}{start_row}"].border = thin_border
    
    start_row += 1

# Insert total row
ws[f"B{start_row}"] = "Total"
ws[f"C{start_row}"] = df["Value"].sum()

# Apply border to the total row
for col in ["A", "B", "C"]:
    ws[f"{col}{start_row}"].border = thin_border

# Update row 10 with the total sum
ws["A10"] = f'Record Keeping Status Break-out (IP Deferral Report) of Participants "Not on GM Pay File" - {df["Value"].sum()}'

# Save the updated Excel file
wb.save(file_path)
